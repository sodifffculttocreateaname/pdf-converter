# -*- coding: utf-8 -*-
"""
测试用例生成脚本
生成PDF工具箱的完整测试用例Excel文件
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def create_test_cases():
    # 创建工作簿
    wb = openpyxl.Workbook()

    # 定义样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def set_column_widths(ws, widths):
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def write_header(ws, row, headers):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

    def write_row(ws, row, data):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = left_align if col > 1 else center_align

    # ==================== 测试用例数据 ====================

    test_cases = {
        '文档转换模块': [
            ['TC-IMG-001', '图片转PDF-单张图片', '图片转PDF', '单张JPG图片转换为PDF',
             '1. 启动应用\n2. 选择"图片转PDF"\n3. 拖入test.jpg\n4. 点击开始转换',
             '生成test.pdf，状态显示成功', ''],
            ['TC-IMG-002', '图片转PDF-多张图片', '图片转PDF', '多张图片合并为一个PDF',
             '1. 选择"图片转PDF"\n2. 添加test1.jpg, test2.png, test3.bmp\n3. 勾选"合并为一个PDF"\n4. 点击开始转换',
             '生成合并后的PDF，包含3张图片', ''],
            ['TC-IMG-003', '图片转PDF-无效文件', '图片转PDF', '添加非图片文件应被过滤',
             '1. 尝试添加test.txt和test.pdf到图片列表',
             '文件被过滤，不添加到列表', ''],
            ['TC-IMG-004', '图片转PDF-DPI设置', '图片转PDF', '测试DPI参数',
             '1. 添加图片文件\n2. 设置DPI为300\n3. 转换',
             '生成的PDF清晰度符合300DPI', ''],
            ['TC-IMG-005', '图片转PDF-质量设置', '图片转PDF', '测试图片质量参数',
             '1. 设置质量为50%\n2. 转换',
             '生成的PDF文件较小', ''],
            ['TC-PDF-IMG-001', 'PDF转图片-基础转换', 'PDF转图片', 'PDF转换为PNG图片',
             '1. 选择"PDF转图片"\n2. 添加test.pdf(3页)\n3. 输出格式选择PNG\n4. DPI设置150\n5. 开始转换',
             '生成3张PNG图片', ''],
            ['TC-PDF-IMG-002', 'PDF转图片-JPEG格式', 'PDF转图片', 'PDF转换为JPEG格式',
             '1. 输出格式选择JPEG\n2. 质量设置为80%\n3. 转换测试',
             '生成JPEG格式图片', ''],
            ['TC-PDF-IMG-003', 'PDF转图片-BMP格式', 'PDF转图片', 'PDF转换为BMP格式',
             '1. 选择BMP格式\n2. 转换',
             '生成BMP格式图片', ''],
            ['TC-TXT-001', 'TXT转PDF-基础转换', 'TXT转PDF', '纯文本文件转换为PDF',
             '1. 选择"TXT转PDF"\n2. 添加test.txt(UTF-8编码)\n3. 字体大小12\n4. 开始转换',
             '生成test.pdf，内容正确显示', ''],
            ['TC-TXT-002', 'TXT转PDF-GBK编码', 'TXT转PDF', 'GBK编码文本转换',
             '1. 添加GBK编码的中文txt文件\n2. 开始转换',
             '中文正确显示，无乱码', ''],
            ['TC-TXT-003', 'TXT转PDF-字体设置', 'TXT转PDF', '测试字体大小设置',
             '1. 字体大小设置为16\n2. 行高设置为2.0\n3. 转换',
             'PDF字体和行高符合设置', ''],
            ['TC-WORD-001', 'Word转PDF-基础转换', 'Word转PDF', 'DOCX文件转换为PDF',
             '1. 选择"Word转PDF"\n2. 添加test.docx\n3. 开始转换',
             '生成test.pdf，格式保持', ''],
            ['TC-WORD-002', 'Word转PDF-DOC格式', 'Word转PDF', 'DOC格式文件转换',
             '1. 添加test.doc文件\n2. 转换',
             '成功转换，格式正确', ''],
            ['TC-EXCEL-001', 'Excel转PDF-基础转换', 'Excel转PDF', 'XLSX文件转换为PDF',
             '1. 选择"Excel转PDF"\n2. 添加test.xlsx\n3. 页面方向选择自动\n4. 开始转换',
             '生成PDF，表格完整显示', ''],
            ['TC-EXCEL-002', 'Excel转PDF-横向', 'Excel转PDF', '横向输出',
             '1. 页面方向选择"横向"\n2. 转换',
             'PDF页面为横向布局', ''],
            ['TC-PPT-001', 'PPT转PDF-基础转换', 'PPT转PDF', 'PPTX文件转换为PDF',
             '1. 选择"PPT转PDF"\n2. 添加test.pptx(10页)\n3. 开始转换',
             '生成10页PDF', ''],
            ['TC-PPT-002', 'PPT转PDF-PPT格式', 'PPT转PDF', 'PPT格式转换',
             '1. 添加test.ppt文件\n2. 转换',
             '成功转换旧版PPT', ''],
            ['TC-WEB-001', '网页转PDF-URL输入', '网页转PDF', '通过URL转换网页',
             '1. 选择"网页转PDF"\n2. 输入https://www.baidu.com\n3. 点击添加\n4. 开始转换',
             '生成网页PDF', ''],
            ['TC-WEB-002', '网页转PDF-HTML文件', '网页转PDF', '本地HTML文件转换',
             '1. 添加local.html文件\n2. 开始转换',
             '生成PDF', ''],
            ['TC-WEB-003', '网页转PDF-页面大小', '网页转PDF', '测试页面大小设置',
             '1. 页面大小选择A3\n2. 转换',
             'PDF为A3尺寸', ''],
            ['TC-DOC-001', '文档转PDF-自动识别', '文档转PDF', '自动识别文档类型',
             '1. 添加多个不同类型文档(txt, docx, xlsx)\n2. 转换',
             '所有文档正确转换为PDF', ''],
        ],
        'PDF基础处理模块': [
            ['TC-MERGE-001', 'PDF合并-基础合并', 'PDF合并', '合并多个PDF文件',
             '1. 选择"PDF合并"\n2. 添加a.pdf, b.pdf, c.pdf\n3. 设置输出文件名\n4. 开始合并',
             '生成合并后的PDF', ''],
            ['TC-MERGE-002', 'PDF合并-调整顺序', 'PDF合并', '调整合并顺序',
             '1. 添加3个PDF\n2. 使用上移/下移调整顺序\n3. 开始合并',
             '按调整后的顺序合并', ''],
            ['TC-MERGE-003', 'PDF合并-清空列表', 'PDF合并', '测试清空功能',
             '1. 添加多个PDF\n2. 点击清空按钮',
             '文件列表被清空', ''],
            ['TC-SPLIT-001', 'PDF拆分-单页拆分', 'PDF拆分', 'PDF拆分为单页',
             '1. 选择"PDF拆分"\n2. 添加test.pdf(5页)\n3. 选择"单页拆分"\n4. 开始拆分',
             '生成5个独立PDF文件', ''],
            ['TC-SPLIT-002', 'PDF拆分-范围拆分', 'PDF拆分', '按页码范围拆分',
             '1. 选择"页码范围拆分"\n2. 输入范围"1-3,5"\n3. 开始拆分',
             '生成指定范围的PDF', ''],
            ['TC-SPLIT-003', 'PDF拆分-奇偶页拆分', 'PDF拆分', '拆分奇偶页',
             '1. 选择"奇偶页拆分"\n2. 添加10页PDF\n3. 开始拆分',
             '生成odd.pdf(5页)和even.pdf(5页)', ''],
            ['TC-SPLIT-004', 'PDF拆分-无效范围', 'PDF拆分', '输入无效页码范围',
             '1. 输入范围"abc"\n2. 开始拆分',
             '提示格式错误', ''],
            ['TC-COMPRESS-001', 'PDF压缩-低压缩', 'PDF压缩', '低压缩级别测试',
             '1. 选择"PDF压缩"\n2. 添加large.pdf(10MB)\n3. 选择"低压缩"\n4. 开始压缩',
             '文件大小减小，质量保持', ''],
            ['TC-COMPRESS-002', 'PDF压缩-高压缩', 'PDF压缩', '高压缩级别测试',
             '1. 选择"高压缩"\n2. 图片质量50%\n3. 压缩测试',
             '文件大幅减小', ''],
            ['TC-COMPRESS-003', 'PDF压缩-质量设置', 'PDF压缩', '测试图片质量设置',
             '1. 自定义图片质量为30%\n2. 压缩',
             '压缩比明显，质量下降', ''],
        ],
        'PDF高级处理模块': [
            ['TC-EXTRACT-001', 'PDF提取图片-基础提取', 'PDF提取图片', '提取PDF中嵌入的图片',
             '1. 选择"PDF提取图片"\n2. 添加含图片的PDF\n3. 最小尺寸100x100\n4. 输出PNG\n5. 开始提取',
             '提取所有符合条件的图片', ''],
            ['TC-EXTRACT-002', 'PDF提取图片-尺寸过滤', 'PDF提取图片', '测试尺寸过滤功能',
             '1. 设置最小尺寸500x500\n2. 提取图片',
             '只提取大于500x500的图片', ''],
            ['TC-EXTRACT-003', 'PDF提取图片-JPG输出', 'PDF提取图片', 'JPG格式输出',
             '1. 输出格式选择JPG\n2. 提取',
             '图片为JPG格式', ''],
            ['TC-ADD-DEL-001', 'PDF增删页-删除页面', 'PDF增删页', '删除指定页面',
             '1. 选择"PDF增删页"\n2. 添加test.pdf(10页)\n3. 选择"删除页面"\n4. 输入"1,3,5-7"\n5. 开始处理',
             '生成删除后6页的PDF', ''],
            ['TC-ADD-DEL-002', 'PDF增删页-插入空白页', 'PDF增删页', '插入空白页',
             '1. 选择"插入空白页"\n2. 输入"1,3"\n3. 开始处理',
             '在第1页和第3页后插入空白页', ''],
            ['TC-ADD-DEL-003', 'PDF增删页-删除全部', 'PDF增删页', '删除所有页面（边界测试）',
             '1. 输入"1-10"删除所有页面',
             '提示错误或生成空PDF', ''],
            ['TC-ROTATE-001', 'PDF旋转-90度旋转', 'PDF旋转页面', '顺时针旋转90度',
             '1. 选择"PDF旋转页面"\n2. 添加test.pdf\n3. 选择"顺时针90°"\n4. 开始旋转',
             '所有页面旋转90度', ''],
            ['TC-ROTATE-002', 'PDF旋转-180度旋转', 'PDF旋转页面', '旋转180度',
             '1. 选择"180°"\n2. 旋转测试',
             '页面旋转180度', ''],
            ['TC-ROTATE-003', 'PDF旋转-270度旋转', 'PDF旋转页面', '顺时针旋转270度',
             '1. 选择"顺时针270°"\n2. 旋转测试',
             '页面旋转270度（逆时针90度）', ''],
            ['TC-ROTATE-004', 'PDF旋转-指定页码', 'PDF旋转页面', '旋转指定页码',
             '1. 页码范围输入"1-3"\n2. 旋转测试',
             '仅第1-3页旋转', ''],
            ['TC-ORGANIZE-001', 'PDF编排-调整顺序', 'PDF编排页面', '调整页面顺序',
             '1. 选择"PDF编排页面"\n2. 加载test.pdf\n3. 拖拽或使用按钮调整顺序\n4. 开始编排',
             '页面顺序按要求调整', ''],
            ['TC-ORGANIZE-002', 'PDF编排-复制页面', 'PDF编排页面', '复制指定页面',
             '1. 选择第2页\n2. 点击"复制"按钮\n3. 编排',
             '第2页被复制', ''],
            ['TC-ORGANIZE-003', 'PDF编排-删除页面', 'PDF编排页面', '删除页面',
             '1. 选择页面\n2. 点击删除',
             '页面被删除', ''],
            ['TC-LONG-IMG-001', 'PDF转长图-基础转换', 'PDF转长图', 'PDF转为长图',
             '1. 选择"PDF转长图"\n2. 添加test.pdf(5页)\n3. DPI 150\n4. 开始转换',
             '生成一张长图包含5页内容', ''],
            ['TC-LONG-IMG-002', 'PDF转长图-带间距', 'PDF转长图', '设置页面间距',
             '1. 页面间距设置为20px\n2. 转换测试',
             '长图有白色间距', ''],
            ['TC-LONG-IMG-003', 'PDF转长图-JPG输出', 'PDF转长图', 'JPG格式输出',
             '1. 输出格式选择JPG\n2. 转换',
             '生成为JPG长图', ''],
            ['TC-GRAY-001', 'PDF转黑白-灰度模式', 'PDF转黑白', '转为灰度PDF',
             '1. 选择"PDF转黑白"\n2. 添加彩色PDF\n3. 选择"灰度模式"\n4. 开始转换',
             '生成灰度PDF', ''],
            ['TC-GRAY-002', 'PDF转黑白-二值化', 'PDF转黑白', '二值化模式测试',
             '1. 选择"二值化模式"\n2. 阈值128\n3. 转换',
             '生成黑白二值PDF', ''],
            ['TC-GRAY-003', 'PDF转黑白-阈值调整', 'PDF转黑白', '调整二值化阈值',
             '1. 阈值设置为200\n2. 转换',
             '二值化效果符合阈值', ''],
            ['TC-PAGE-NUM-001', 'PDF添加页码-居中', 'PDF添加页码', '底部居中添加页码',
             '1. 选择"PDF添加页码"\n2. 添加test.pdf\n3. 位置选择"底部居中"\n4. 格式"1,2,3..."\n5. 开始添加',
             '每页底部显示页码', ''],
            ['TC-PAGE-NUM-002', 'PDF添加页码-带总数', 'PDF添加页码', '显示页码和总数',
             '1. 格式选择"第1页/共N页"\n2. 添加',
             '显示如"第1页/共5页"', ''],
            ['TC-PAGE-NUM-003', 'PDF添加页码-左下角', 'PDF添加页码', '左下角位置',
             '1. 位置选择"左下角"\n2. 添加',
             '页码显示在左下角', ''],
            ['TC-PAGE-NUM-004', 'PDF添加页码-字体大小', 'PDF添加页码', '设置字体大小',
             '1. 字体大小设置为16\n2. 添加',
             '页码字体变大', ''],
            ['TC-CROP-001', 'PDF分割裁剪-裁剪边距', 'PDF分割裁剪', '裁剪页面边距',
             '1. 选择"PDF分割裁剪"\n2. 添加PDF\n3. 选择"裁剪模式"\n4. 设置边距(10,10,10,10)\n5. 开始处理',
             '页面边距被裁剪', ''],
            ['TC-CROP-002', 'PDF分割裁剪-水平分割', 'PDF分割裁剪', '水平分割页面',
             '1. 选择"分割模式"\n2. 选择"水平分割"\n3. 处理',
             '每页分为上下两部分', ''],
            ['TC-CROP-003', 'PDF分割裁剪-垂直分割', 'PDF分割裁剪', '垂直分割页面',
             '1. 选择"分割模式"\n2. 选择"垂直分割"\n3. 处理',
             '每页分为左右两部分', ''],
            ['TC-PAGE-MERGE-001', 'PDF页面合并-2合1', 'PDF页面合并', '2页合并为1页',
             '1. 选择"PDF页面合并"\n2. 添加test.pdf(6页)\n3. 选择"2合1"\n4. 开始合并',
             '生成3页PDF，每页含2个原页面', ''],
            ['TC-PAGE-MERGE-002', 'PDF页面合并-4合1', 'PDF页面合并', '4页合并为1页',
             '1. 选择"PDF页面合并"\n2. 添加test.pdf(8页)\n3. 选择"4合1"\n4. 开始合并',
             '生成2页PDF，每页含4个原页面', ''],
            ['TC-PAGE-MERGE-003', 'PDF页面合并-带边框', 'PDF页面合并', '显示边框测试',
             '1. 勾选"显示边框"\n2. 合并测试',
             '每个小页面有边框', ''],
            ['TC-PAGE-MERGE-004', 'PDF页面合并-间距设置', 'PDF页面合并', '设置页面间距',
             '1. 页面间距设置为10px\n2. 合并',
             '小页面之间有间距', ''],
        ],
        'PDF安全与水印模块': [
            ['TC-WATERMARK-001', 'PDF加水印-文字水印', 'PDF加水印', '添加文字水印',
             '1. 选择"PDF加水印"\n2. 添加PDF\n3. 选择"文字水印"\n4. 输入"CONFIDENTIAL"\n5. 透明度30%\n6. 开始添加',
             '每页显示半透明水印', ''],
            ['TC-WATERMARK-002', 'PDF加水印-图片水印', 'PDF加水印', '添加图片水印',
             '1. 选择"图片水印"\n2. 选择logo.png\n3. 位置"居中"\n4. 添加',
             '每页显示图片水印', ''],
            ['TC-WATERMARK-003', 'PDF加水印-平铺模式', 'PDF加水印', '平铺水印测试',
             '1. 位置选择"平铺"\n2. 添加水印',
             '水印平铺覆盖整个页面', ''],
            ['TC-WATERMARK-004', 'PDF加水印-旋转角度', 'PDF加水印', '设置旋转角度',
             '1. 旋转角度设置为45°\n2. 添加水印',
             '水印以45度角显示', ''],
            ['TC-WATERMARK-005', 'PDF加水印-透明度', 'PDF加水印', '测试透明度设置',
             '1. 透明度设置为50%\n2. 添加水印',
             '水印透明度正确', ''],
            ['TC-REMOVE-WM-001', 'PDF去水印-基础去除', 'PDF去水印', '去除透明水印',
             '1. 选择"PDF去水印"\n2. 添加带水印的PDF\n3. 阈值0.3\n4. 开始处理',
             '透明水印被移除', ''],
            ['TC-REMOVE-WM-002', 'PDF去水印-高阈值', 'PDF去水印', '高阈值测试',
             '1. 阈值设置为0.5\n2. 去水印',
             '更多内容被移除', ''],
            ['TC-ENCRYPT-001', 'PDF加密-设置打开密码', 'PDF加密', '设置打开密码',
             '1. 选择"PDF加密"\n2. 添加PDF\n3. 输入打开密码"test123"\n4. 确认密码\n5. 开始加密',
             '生成加密PDF，打开需密码', ''],
            ['TC-ENCRYPT-002', 'PDF加密-权限设置', 'PDF加密', '设置编辑权限',
             '1. 设置打开密码\n2. 取消"允许修改"\n3. 加密',
             'PDF不可编辑', ''],
            ['TC-ENCRYPT-003', 'PDF加密-密码确认', 'PDF加密', '密码不一致测试',
             '1. 输入打开密码"123"\n2. 确认密码"456"\n3. 尝试加密',
             '提示密码不一致', ''],
            ['TC-ENCRYPT-004', 'PDF加密-权限密码', 'PDF加密', '设置权限密码',
             '1. 设置打开密码"open"\n2. 设置权限密码"admin"\n3. 加密',
             '两个密码都设置成功', ''],
            ['TC-ENCRYPT-005', 'PDF加密-打印权限', 'PDF加密', '设置打印权限',
             '1. 设置密码\n2. 取消"允许打印"\n3. 加密',
             'PDF不可打印', ''],
            ['TC-INVOICE-001', '发票合并-基础合并', '发票合并', '合并多张发票',
             '1. 选择"发票合并"\n2. 添加invoice1.pdf, invoice2.pdf, invoice3.pdf\n3. 输出文件名"merged.pdf"\n4. 开始合并',
             '生成合并后的发票PDF', ''],
            ['TC-INVOICE-002', '发票合并-调整顺序', '发票合并', '调整发票顺序',
             '1. 添加3张发票\n2. 使用上移/下移调整顺序\n3. 合并',
             '按调整顺序合并', ''],
            ['TC-INVOICE-003', '发票合并-清空列表', '发票合并', '测试清空功能',
             '1. 添加发票文件\n2. 点击清空',
             '列表被清空', ''],
        ],
        '边界与异常测试': [
            ['TC-EDGE-001', '空文件列表', '所有模块', '未添加文件时点击开始',
             '1. 不添加任何文件\n2. 点击开始按钮',
             '提示"请先添加文件"', ''],
            ['TC-EDGE-002', '大文件处理', 'PDF相关模块', '处理大PDF文件(100MB+)',
             '1. 添加100MB的PDF文件\n2. 执行转换',
             '成功处理，无内存溢出', ''],
            ['TC-EDGE-003', '多文件批量处理', '所有模块', '批量处理100个文件',
             '1. 添加100个文件\n2. 开始处理',
             '全部成功处理', ''],
            ['TC-EDGE-004', '取消操作', '所有模块', '中途取消任务',
             '1. 开始处理\n2. 处理过程中点击取消',
             '任务被取消，已处理的部分保留', ''],
            ['TC-EDGE-005', '重复添加文件', '文件选择', '重复添加同一文件',
             '1. 添加test.pdf\n2. 再次添加test.pdf',
             '文件不重复添加', ''],
            ['TC-EDGE-006', '输出目录不存在', '所有模块', '输出目录不存在时自动创建',
             '1. 设置不存在的输出目录\n2. 开始处理',
             '目录自动创建，处理成功', ''],
            ['TC-EDGE-007', '无效PDF文件', 'PDF相关模块', '处理损坏的PDF',
             '1. 添加损坏的PDF文件\n2. 尝试处理',
             '显示错误信息，不崩溃', ''],
            ['TC-EDGE-008', '网络超时', '网页转PDF', '访问超时网页',
             '1. 输入无法访问的URL\n2. 超时设置30秒\n3. 转换',
             '超时后显示错误', ''],
            ['TC-EDGE-009', '编码异常', 'TXT转PDF', '处理特殊编码文件',
             '1. 添加非标准编码的txt文件\n2. 转换',
             '尝试多种编码或提示错误', ''],
            ['TC-EDGE-010', '退出确认', '应用', '处理中退出应用',
             '1. 开始处理任务\n2. 关闭窗口',
             '弹出确认对话框', ''],
            ['TC-EDGE-011', '添加文件夹', '支持文件夹的模块', '拖拽文件夹添加文件',
             '1. 拖拽包含PDF文件的文件夹\n2. 观察文件列表',
             '文件夹内的PDF被添加', ''],
            ['TC-EDGE-012', '右键移除文件', '文件列表', '右键删除文件',
             '1. 右键点击文件列表项\n2. 选择"移除"',
             '文件被移除', ''],
            ['TC-EDGE-013', '进度条显示', '所有模块', '处理过程中进度更新',
             '1. 开始处理多文件\n2. 观察进度条',
             '进度条正确更新', ''],
            ['TC-EDGE-014', '状态栏提示', '主界面', '选择功能后状态栏更新',
             '1. 点击不同功能项\n2. 观察状态栏',
             '状态栏显示当前功能', ''],
            ['TC-EDGE-015', '按钮状态', '所有模块', '处理中按钮状态',
             '1. 开始处理\n2. 观察按钮状态',
             '开始按钮禁用，取消按钮启用', ''],
        ],
        '界面交互测试': [
            ['TC-UI-001', '页面切换', '主界面', '切换不同功能页面',
             '1. 点击左侧不同功能项\n2. 观察右侧页面变化',
             '页面正确切换', ''],
            ['TC-UI-002', '拖拽上传', '所有模块', '拖拽文件到上传区域',
             '1. 拖拽文件到虚线框区域',
             '文件被正确添加', ''],
            ['TC-UI-003', '点击上传区域', '所有模块', '点击上传区域选择文件',
             '1. 点击拖拽上传区域\n2. 在弹窗中选择文件',
             '文件被添加', ''],
            ['TC-UI-004', '文件大小显示', '文件列表', '显示文件大小',
             '1. 添加文件\n2. 查看大小列',
             '显示格式化的大小(如1.5MB)', ''],
            ['TC-UI-005', '文件状态显示', '文件列表', '处理状态更新',
             '1. 开始处理\n2. 观察状态列变化',
             '状态正确显示(待处理->处理中->已完成)', ''],
            ['TC-UI-006', '设置持久化', '设置项', '设置项保存',
             '1. 修改设置(如DPI)\n2. 切换页面\n3. 返回',
             '设置保持不变', ''],
            ['TC-UI-007', '窗口大小调整', '主界面', '调整窗口大小',
             '1. 拖拽窗口边缘调整大小\n2. 观察布局',
             '布局正确自适应', ''],
            ['TC-UI-008', '文件列表滚动', '文件列表', '多文件滚动',
             '1. 添加20个文件\n2. 滚动查看',
             '滚动条正常，显示正确', ''],
            ['TC-UI-009', '错误提示', '所有模块', '处理失败显示错误',
             '1. 添加无效文件\n2. 处理',
             '显示错误信息', ''],
            ['TC-UI-010', '成功提示', '所有模块', '处理成功显示结果',
             '1. 正常处理完成',
             '显示成功消息框', ''],
        ],
    }

    # 创建各模块工作表
    for sheet_name, cases in test_cases.items():
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel限制工作表名31字符

        # 写入表头
        headers = ['用例编号', '用例名称', '所属模块', '用例描述', '测试步骤', '预期结果', '实际结果', '测试状态', '测试日期', '测试人员']
        write_header(ws, 1, headers)

        # 写入数据
        for i, case in enumerate(cases, 2):
            # 添加空的实际结果、测试状态、测试日期、测试人员列供填写
            row_data = case + ['', '', '', '']
            write_row(ws, i, row_data)

        # 设置列宽
        set_column_widths(ws, [15, 25, 15, 30, 50, 30, 30, 12, 15, 15])

    # 删除默认的Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # 创建汇总页
    ws_summary = wb.create_sheet(title='测试汇总', index=0)

    # 标题
    ws_summary.cell(row=1, column=1, value='PDF工具箱测试用例汇总')
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:E1')

    # 汇总表头
    summary_headers = ['模块分类', '用例数量', '优先级', '负责人', '备注']
    write_header(ws_summary, 3, summary_headers)

    summary_data = [
        ['文档转换模块', len(test_cases['文档转换模块']), '高', '', '图片/PDF/TXT/Word/Excel/PPT/网页转换'],
        ['PDF基础处理模块', len(test_cases['PDF基础处理模块']), '高', '', '合并/拆分/压缩'],
        ['PDF高级处理模块', len(test_cases['PDF高级处理模块']), '中', '', '提取/旋转/编排/长图/黑白/页码/裁剪/合并'],
        ['PDF安全与水印模块', len(test_cases['PDF安全与水印模块']), '中', '', '水印/加密/发票合并'],
        ['边界与异常测试', len(test_cases['边界与异常测试']), '高', '', '异常处理和边界条件'],
        ['界面交互测试', len(test_cases['界面交互测试']), '中', '', 'UI交互测试'],
    ]

    for i, row in enumerate(summary_data, 4):
        write_row(ws_summary, i, row)

    # 添加总计行
    total_row = len(summary_data) + 4
    ws_summary.cell(row=total_row, column=1, value='总计')
    ws_summary.cell(row=total_row, column=2, value=sum(r[1] for r in summary_data))
    ws_summary.cell(row=total_row, column=1).font = Font(bold=True)
    ws_summary.cell(row=total_row, column=2).font = Font(bold=True)
    for col in range(1, 6):
        ws_summary.cell(row=total_row, column=col).border = thin_border

    set_column_widths(ws_summary, [20, 15, 15, 15, 40])

    # 添加测试说明
    ws_summary.cell(row=total_row + 2, column=1, value='测试说明:')
    ws_summary.cell(row=total_row + 2, column=1).font = Font(bold=True)

    notes = [
        '1. 每个测试用例需填写实际结果、测试状态(通过/失败)、测试日期和测试人员',
        '2. 优先级：高-核心功能，中-次要功能，低-辅助功能',
        '3. 测试环境：Windows 11, Python 3.x, PySide6',
        '4. 如发现问题，请在"实际结果"列详细描述问题现象',
        f'5. 用例总数: {sum(len(cases) for cases in test_cases.values())}',
    ]

    for i, note in enumerate(notes, total_row + 3):
        ws_summary.cell(row=i, column=1, value=note)

    # 保存文件
    output_path = 'testPdfTool.xlsx'
    wb.save(output_path)
    print(f'测试用例文件已生成: {output_path}')
    print(f'总用例数: {sum(len(cases) for cases in test_cases.values())}')
    return output_path

if __name__ == '__main__':
    create_test_cases()