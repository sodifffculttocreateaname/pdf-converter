# -*- coding: utf-8 -*-
"""
PDF转Excel转换器
"""
from pathlib import Path
from typing import List

import pdfplumber
from openpyxl import Workbook

from config.constants import ConversionStatus
from core.base_converter import BaseConverter, ConversionProgress, ConversionResult
from utils.file_utils import ensure_dir
from utils.logger import get_logger

logger = get_logger(__name__)


class PdfToExcelConverter(BaseConverter):
    """PDF转Excel转换器"""

    name = "PDF转Excel"
    description = "将PDF中的表格提取为Excel"
    supported_input_formats = ["pdf"]
    supported_output_formats = ["xlsx"]

    def __init__(self):
        super().__init__()
        self._extract_all_pages = True
        self._pages_to_extract = []  # 要提取的页码列表

    def set_pages(self, pages: List[int]):
        """
        设置要提取的页码

        Args:
            pages: 页码列表（从1开始）
        """
        self._extract_all_pages = False
        self._pages_to_extract = pages

    def convert(self, input_files: List[Path], output_dir: Path, **kwargs) -> List[ConversionResult]:
        """
        执行PDF转Excel转换

        Args:
            input_files: 输入PDF文件列表
            output_dir: 输出目录
            **kwargs: 额外参数 (pages)

        Returns:
            转换结果列表
        """
        results = []

        # 应用参数
        pages = kwargs.get('pages', None)
        if pages:
            self.set_pages(pages)

        # 确保输出目录存在
        ensure_dir(output_dir)

        total_files = len(input_files)

        for i, input_file in enumerate(input_files):
            if self._check_cancelled():
                break

            # 报告进度
            progress = ConversionProgress(
                current=i + 1,
                total=total_files,
                current_file=input_file.name,
                message=f"正在提取: {input_file.name}"
            )
            self._report_progress(progress)

            result = self._convert_pdf_to_excel(input_file, output_dir)
            results.append(result)

        # 报告完成进度
        if not self._check_cancelled():
            progress = ConversionProgress(
                current=total_files,
                total=total_files,
                message="处理完成"
            )
            self._report_progress(progress)

        return results

    def _convert_pdf_to_excel(self, input_file: Path, output_dir: Path) -> ConversionResult:
        """
        转换单个PDF为Excel

        Args:
            input_file: 输入PDF文件
            output_dir: 输出目录

        Returns:
            转换结果
        """
        result = ConversionResult(input_file=input_file)

        try:
            # 生成输出文件名
            output_filename = self.get_output_filename(input_file, 'xlsx')
            output_path = output_dir / output_filename

            # 创建Excel工作簿
            wb = Workbook()
            # 删除默认创建的工作表
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

            tables_found = 0

            with pdfplumber.open(str(input_file)) as pdf:
                total_pages = len(pdf.pages)

                # 确定要处理的页面
                if self._extract_all_pages:
                    pages_to_process = range(total_pages)
                else:
                    pages_to_process = [p - 1 for p in self._pages_to_extract if 0 < p <= total_pages]

                for page_idx in pages_to_process:
                    if self._check_cancelled():
                        break

                    page = pdf.pages[page_idx]
                    page_num = page_idx + 1

                    # 报告进度
                    progress = ConversionProgress(
                        current=page_idx + 1,
                        total=total_pages,
                        current_file=input_file.name,
                        message=f"正在处理第 {page_num} 页"
                    )
                    self._report_progress(progress)

                    # 提取表格
                    tables = page.extract_tables()

                    if tables:
                        for table_idx, table in enumerate(tables):
                            if not table:
                                continue

                            # 创建工作表
                            sheet_name = f"第{page_num}页_表{table_idx + 1}"
                            if len(sheet_name) > 31:  # Excel工作表名称限制
                                sheet_name = f"P{page_num}_T{table_idx + 1}"

                            ws = wb.create_sheet(title=sheet_name)

                            # 写入数据
                            for row_idx, row in enumerate(table, 1):
                                for col_idx, cell in enumerate(row, 1):
                                    if cell is not None:
                                        ws.cell(row=row_idx, column=col_idx, value=str(cell))

                            tables_found += 1

            if self._check_cancelled():
                result.status = ConversionStatus.CANCELLED
                if output_path.exists():
                    output_path.unlink()
                return result

            if tables_found == 0:
                result.status = ConversionStatus.FAILED
                result.error = "未在PDF中找到表格"
                return result

            # 保存Excel文件
            wb.save(str(output_path))

            result.output_file = output_path
            result.status = ConversionStatus.COMPLETED
            result.message = f"成功提取 {tables_found} 个表格"

            logger.info(f"PDF转Excel成功: {input_file.name}, 提取 {tables_found} 个表格")

        except Exception as e:
            result.status = ConversionStatus.FAILED
            result.error = str(e)
            logger.error(f"PDF转Excel失败: {input_file.name} - {e}")

        return result

    def validate_input(self, file_path: Path) -> bool:
        """
        验证输入文件

        Args:
            file_path: 文件路径

        Returns:
            如果文件有效返回 True
        """
        if not file_path.exists():
            return False

        return file_path.suffix.lower() == '.pdf'