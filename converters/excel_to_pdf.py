# -*- coding: utf-8 -*-
"""
Excel转PDF转换器
支持多种转换方式:
1. Windows COM (仅Windows，需要安装Microsoft Excel)
2. openpyxl + reportlab (跨平台，但格式可能不完全准确)
"""
import platform
from pathlib import Path
from typing import List

from config.constants import ConversionStatus
from core.base_converter import BaseConverter, ConversionProgress, ConversionResult
from utils.file_utils import ensure_dir
from utils.logger import get_logger

logger = get_logger(__name__)


class ExcelToPdfConverter(BaseConverter):
    """Excel转PDF转换器"""

    name = "Excel转PDF"
    description = "将Excel表格转换为PDF"
    supported_input_formats = ["xls", "xlsx"]
    supported_output_formats = ["pdf"]

    def __init__(self):
        super().__init__()
        self._use_com = self._check_com_available()

    def _check_com_available(self) -> bool:
        """检查COM是否可用（仅Windows，且需要真正的Microsoft Excel）"""
        if platform.system() != 'Windows':
            return False

        try:
            import win32com.client
            import pythoncom

            # 尝试创建Excel应用实例
            pythoncom.CoInitialize()
            try:
                excel = win32com.client.Dispatch("Excel.Application")
                excel_path = excel.Path.lower()
                excel.Quit()

                # 检查是否是Microsoft Excel（而不是WPS）
                # WPS的路径通常包含"wps"、"kingsoft"等字样
                if 'wps' in excel_path or 'kingsoft' in excel_path:
                    logger.info(f"检测到WPS Office，将使用openpyxl方法进行转换")
                    return False

                # 检查Excel版本，确保是真正的Microsoft Excel
                if 'microsoft' in excel_path or 'office' in excel_path:
                    logger.info(f"检测到Microsoft Excel，路径: {excel_path}")
                    return True

                # 默认情况下，如果不是WPS，尝试使用COM
                return True
            finally:
                pythoncom.CoUninitialize()

        except ImportError:
            return False
        except Exception as e:
            logger.warning(f"检查COM可用性时出错: {e}")
            return False

    def convert(self, input_files: List[Path], output_dir: Path, **kwargs) -> List[ConversionResult]:
        """
        执行Excel转PDF转换

        Args:
            input_files: 输入Excel文件列表
            output_dir: 输出目录
            **kwargs: 额外参数

        Returns:
            转换结果列表
        """
        results = []

        # 确保输出目录存在
        ensure_dir(output_dir)

        # 选择转换方法
        if self._use_com:
            results = self._convert_with_com(input_files, output_dir)
        else:
            results = self._convert_with_openpyxl(input_files, output_dir)

        return results

    def _convert_with_com(self, input_files: List[Path], output_dir: Path) -> List[ConversionResult]:
        """
        使用Windows COM转换（需要安装Microsoft Excel）

        Args:
            input_files: 输入文件列表
            output_dir: 输出目录

        Returns:
            转换结果列表
        """
        import win32com.client
        import pythoncom

        results = []
        excel_app = None

        try:
            # 初始化COM
            pythoncom.CoInitialize()

            # 创建Excel应用
            excel_app = win32com.client.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            excel_app.ScreenUpdating = False

            total_files = len(input_files)

            for i, input_file in enumerate(input_files):
                if self._check_cancelled():
                    break

                # 报告进度
                progress = ConversionProgress(
                    current=i + 1,
                    total=total_files,
                    current_file=input_file.name,
                    message=f"正在转换: {input_file.name}"
                )
                self._report_progress(progress)

                result = ConversionResult(input_file=input_file)

                workbook = None
                try:
                    # 确保文件存在且可访问
                    if not input_file.exists():
                        raise FileNotFoundError(f"文件不存在: {input_file}")

                    abs_path = str(input_file.absolute())
                    logger.info(f"正在打开Excel文件: {abs_path}")

                    # 打开工作簿 - 使用更完整的参数
                    workbook = excel_app.Workbooks.Open(
                        Filename=abs_path,
                        ReadOnly=True,
                        UpdateLinks=0,  # 不更新外部链接
                        IgnoreReadOnlyRecommended=True
                    )

                    if workbook is None:
                        raise RuntimeError("无法打开工作簿，Workbooks.Open返回None")

                    # 遍历所有工作表，自动调整列宽并设置页面格式
                    for sheet in workbook.Worksheets:
                        # 检查工作表是否有数据
                        if sheet.UsedRange is None or sheet.UsedRange.Rows.Count == 0:
                            logger.warning(f"工作表 {sheet.Name} 为空，跳过")
                            continue

                        # 先自动调整所有列宽以适应内容
                        try:
                            sheet.Columns.AutoFit()
                        except Exception as e:
                            logger.warning(f"自动调整列宽失败: {e}")

                        # 自动调整行高以适应内容
                        try:
                            sheet.Rows.AutoFit()
                        except Exception as e:
                            logger.warning(f"自动调整行高失败: {e}")

                        # 设置页面方向为横向（适合宽表格）
                        try:
                            sheet.PageSetup.Orientation = 2  # 2 = xlLandscape
                        except Exception as e:
                            logger.warning(f"设置页面方向失败: {e}")

                        # 设置纸张大小为A4
                        try:
                            sheet.PageSetup.PaperSize = 9  # 9 = xlPaperA4
                        except Exception as e:
                            logger.warning(f"设置纸张大小失败: {e}")

                        # 设置打印区域为已使用的区域
                        try:
                            if sheet.UsedRange.Address:
                                sheet.PageSetup.PrintArea = sheet.UsedRange.Address
                        except Exception as e:
                            logger.warning(f"设置打印区域失败: {e}")

                        # 计算需要的缩放比例（根据内容宽度）
                        try:
                            # 获取已使用区域的宽度（以磅为单位）
                            used_width = sheet.UsedRange.Width
                            # A4横向可打印区域约为 842 - 边距(约70) = 772 磅
                            page_width = 772

                            if used_width > page_width:
                                # 内容太宽，需要缩放
                                zoom_percent = int((page_width / used_width) * 100)
                                # 限制缩放范围在10%-400%（Excel的限制）
                                zoom_percent = max(min(zoom_percent, 400), 10)
                                sheet.PageSetup.Zoom = zoom_percent
                                logger.info(f"工作表 {sheet.Name}: 内容宽度{used_width:.0f}pt，设置缩放为{zoom_percent}%")
                            else:
                                # 内容不宽，按100%显示
                                sheet.PageSetup.Zoom = 100
                        except Exception as e:
                            logger.warning(f"设置缩放比例失败: {e}")

                    # 生成输出文件名
                    output_filename = self.get_output_filename(input_file, 'pdf')
                    output_path = output_dir / output_filename

                    # 导出为PDF
                    # xlTypePDF = 0
                    workbook.ExportAsFixedFormat(0, str(output_path.absolute()))
                    workbook.Close(SaveChanges=False)
                    workbook = None

                    result.output_file = output_path
                    result.status = ConversionStatus.COMPLETED
                    result.message = "转换成功"

                    logger.info(f"Excel转PDF成功: {input_file.name}")

                except Exception as e:
                    result.status = ConversionStatus.FAILED
                    result.error = str(e)
                    logger.error(f"Excel转PDF失败: {input_file.name} - {e}")

                    # 确保关闭工作簿
                    if workbook:
                        try:
                            workbook.Close(SaveChanges=False)
                        except:
                            pass
                        workbook = None

                results.append(result)

        except Exception as e:
            logger.error(f"COM转换失败: {e}")
            return self._convert_with_openpyxl(input_files, output_dir)

        finally:
            # 关闭Excel应用
            if excel_app:
                try:
                    excel_app.Quit()
                except Exception:
                    pass

            # 清理COM
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

        # 报告完成进度
        if not self._check_cancelled():
            progress = ConversionProgress(
                current=len(input_files),
                total=len(input_files),
                message="处理完成"
            )
            self._report_progress(progress)

        return results

    def _convert_with_openpyxl(self, input_files: List[Path], output_dir: Path) -> List[ConversionResult]:
        """
        使用openpyxl和reportlab转换（跨平台，格式可能不完全准确）

        Args:
            input_files: 输入文件列表
            output_dir: 输出目录

        Returns:
            转换结果列表
        """
        from openpyxl import load_workbook
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        results = []

        # 注册字体
        self._register_font()

        total_files = len(input_files)

        for i, input_file in enumerate(input_files):
            if self._check_cancelled():
                break

            # 报告进度
            progress = ConversionProgress(
                current=i + 1,
                total=total_files,
                current_file=input_file.name,
                message=f"正在转换: {input_file.name}"
            )
            self._report_progress(progress)

            result = ConversionResult(input_file=input_file)

            try:
                # 读取Excel
                wb = load_workbook(str(input_file), read_only=True, data_only=True)

                # 生成输出文件名
                output_filename = self.get_output_filename(input_file, 'pdf')
                output_path = output_dir / output_filename

                # 创建PDF - 使用横向A4以获得更宽的页面
                pdf_doc = SimpleDocTemplate(
                    str(output_path),
                    pagesize=landscape(A4),
                    leftMargin=1*cm,
                    rightMargin=1*cm,
                    topMargin=1*cm,
                    bottomMargin=1*cm
                )

                # 创建样式
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'TitleStyle',
                    parent=styles['Heading2'],
                    fontName=self._font_name,
                    fontSize=14,
                )

                # 创建单元格文本样式 - 支持自动换行
                cell_style = ParagraphStyle(
                    'CellStyle',
                    parent=styles['Normal'],
                    fontName=self._font_name,
                    fontSize=9,
                    leading=12,  # 行高
                    wordWrap='CJK',  # 中文自动换行
                )

                # 创建表头样式
                header_style = ParagraphStyle(
                    'HeaderStyle',
                    parent=styles['Normal'],
                    fontName=self._font_name,
                    fontSize=9,
                    leading=12,
                    textColor=colors.whitesmoke,
                    wordWrap='CJK',
                )

                # 构建内容
                story = []

                for sheet_num, sheet_name in enumerate(wb.sheetnames, 1):
                    if self._check_cancelled():
                        break

                    sheet = wb[sheet_name]

                    # 添加工作表标题
                    story.append(Paragraph(f"工作表: {sheet_name}", title_style))
                    story.append(Spacer(1, 12))

                    # 获取数据范围
                    max_row = min(sheet.max_row, 100)  # 限制行数
                    max_col = min(sheet.max_column, 20)  # 限制列数

                    # 提取数据
                    data = []
                    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                        row_data = []
                        for cell in row:
                            value = cell.value
                            if value is None:
                                row_data.append("")
                            else:
                                row_data.append(str(value))
                        data.append(row_data)

                    if data:
                        # 计算自适应列宽
                        # 获取页面可用宽度（A4横向约27.7cm，减去边距）
                        available_width = 27.7*cm - 2*cm  # 总宽度减去左右边距

                        # 计算每列的最大内容长度
                        col_max_lengths = []
                        for col_idx in range(len(data[0])):
                            max_len = 0
                            for row in data:
                                if col_idx < len(row):
                                    cell_text = str(row[col_idx]) if row[col_idx] else ""
                                    # 计算中文字符和英文字符的混合长度
                                    # 中文字符算2.0个单位宽度，英文字符算1.0个单位（中文字更宽）
                                    text_width = sum(2.0 if ord(c) > 127 else 1.0 for c in cell_text)
                                    max_len = max(max_len, text_width)
                            col_max_lengths.append(max_len)

                        # 根据内容长度计算列宽
                        # 最小列宽2cm，字符宽度0.35cm
                        base_width = 2.0*cm
                        char_width = 0.35*cm

                        col_widths = []
                        total_content_width = 0
                        for max_len in col_max_lengths:
                            # 根据内容计算宽度，不设上限，优先保证内容完整显示
                            width = max(base_width, max_len * char_width)
                            col_widths.append(width)
                            total_content_width += width

                        # 如果总宽度超过页面宽度，按比例缩放所有列
                        if total_content_width > available_width:
                            scale_factor = available_width / total_content_width
                            col_widths = [w * scale_factor for w in col_widths]

                        # 将数据转换为Paragraph对象以支持自动换行
                        # 第一行是表头，使用header_style
                        table_data = []
                        for row_idx, row in enumerate(data):
                            new_row = []
                            for col_idx, cell in enumerate(row):
                                cell_text = str(cell) if cell is not None else ""
                                # 对超长文本进行处理，避免不换行的问题
                                # Paragraph会自动处理换行，不需要额外处理
                                if row_idx == 0:
                                    # 表头
                                    new_row.append(Paragraph(cell_text, header_style))
                                else:
                                    # 数据行
                                    new_row.append(Paragraph(cell_text, cell_style))
                            table_data.append(new_row)

                        # 创建表格样式
                        table_style = TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                            ('TOPPADDING', (0, 0), (-1, -1), 6),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                            ('LEFTPADDING', (0, 0), (-1, -1), 4),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                        ])

                        # 创建表格
                        table = Table(table_data, colWidths=col_widths, repeatRows=1)
                        table.setStyle(table_style)

                        story.append(table)

                    # 添加分页（除了最后一个工作表）
                    if sheet_num < len(wb.sheetnames):
                        story.append(PageBreak())

                wb.close()

                if self._check_cancelled():
                    result.status = ConversionStatus.CANCELLED
                    if output_path.exists():
                        output_path.unlink()
                    results.append(result)
                    continue

                # 生成PDF
                pdf_doc.build(story)

                result.output_file = output_path
                result.status = ConversionStatus.COMPLETED
                result.message = f"转换成功 ({len(wb.sheetnames)} 个工作表)"

                logger.info(f"Excel转PDF成功: {input_file.name}")

            except Exception as e:
                result.status = ConversionStatus.FAILED
                result.error = str(e)
                logger.error(f"Excel转PDF失败: {input_file.name} - {e}")

            results.append(result)

        # 报告完成进度
        if not self._check_cancelled():
            progress = ConversionProgress(
                current=total_files,
                total=total_files,
                message="处理完成"
            )
            self._report_progress(progress)

        return results

    _font_registered = False
    _font_name = "SimSun"

    def _register_font(self):
        """注册中文字体"""
        if self._font_registered:
            return

        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        try:
            font_paths = [
                r"C:\Windows\Fonts\simhei.ttf",
                r"C:\Windows\Fonts\simsun.ttc",
                r"C:\Windows\Fonts\msyh.ttc",
                "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
                "/System/Library/Fonts/PingFang.ttc",
            ]

            for font_path in font_paths:
                if Path(font_path).exists():
                    try:
                        pdfmetrics.registerFont(TTFont(self._font_name, font_path))
                        self._font_registered = True
                        logger.info(f"成功注册字体: {font_path}")
                        return
                    except Exception:
                        continue

            self._font_name = "Helvetica"

        except Exception as e:
            logger.warning(f"注册字体失败: {e}")
            self._font_name = "Helvetica"

        self._font_registered = True

    def validate_input(self, file_path: Path) -> bool:
        """
        验证输入文件

        Args:
            file_path: 文件路径

        Returns:
            如果文件有效返回 True
        """
        if not file_path.exists():
            return False

        return file_path.suffix.lower() in ['.xls', '.xlsx']